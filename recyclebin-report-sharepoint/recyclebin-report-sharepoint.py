import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

# =========================
# CONFIG
# =========================

TOKEN = ""

SITE_ID = (
    "accor.sharepoint.com,"
    "a492f6be-ff59-436a-a787-c1b229777590,"
    "3e279574-1667-4526-aec6-9302a7b09340"
)

NEXT_URL = (
    f"https://graph.microsoft.com/beta/sites/{SITE_ID}/recycleBin/items"
    "?$orderby=deletedDateTime desc"
    "&$select=deletedFromLocation,name,size,deletedBy,deletedDateTime"
    "&$top=15000"
)

headers = {
    "Authorization": f"Bearer {TOKEN}"
}

# =========================
# HELPERS
# =========================

def get_dept(path: str) -> str:
    if not path:
        return "UNKNOWN"
    parts = path.split("/")
    return parts[2] if len(parts) >= 3 else "UNKNOWN"


def format_size(size_bytes: int) -> str:
    kb = size_bytes / 1024
    mb = kb / 1024
    gb = mb / 1024

    if gb >= 1:
        return f"{gb:.2f} GB"
    elif mb >= 1:
        return f"{mb:.2f} MB"
    else:
        return f"{kb:.2f} KB"


def to_vn_time(utc_str: str) -> str:
    if not utc_str:
        return ""
    dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00"))
    return (dt + timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S")


# =========================
# STORAGE
# =========================

dept_total_size = {}
dept_file_count = {}

all_rows = []

page_count = 0
total_items = 0

# MAX_PAGES = 1  # TEST MODE

# =========================
# FETCH DATA
# =========================

while NEXT_URL:
    page_count += 1
    print(f"Page {page_count}")

    res = requests.get(NEXT_URL, headers=headers, timeout=60)
    res.raise_for_status()

    data = res.json()
    items = data.get("value", [])

    for item in items:
        path = item.get("deletedFromLocation", "")
        dept = get_dept(path)

        size_bytes = int(item.get("size", 0) or 0)

        deleted_by = (
            item.get("deletedBy", {})
                .get("user", {})
                .get("displayName", "")
        )

        vn_time = to_vn_time(item.get("deletedDateTime", ""))

        # dashboard stats
        dept_total_size[dept] = dept_total_size.get(dept, 0) + size_bytes
        dept_file_count[dept] = dept_file_count.get(dept, 0) + 1

        # store unified rows (NO UTC COLUMN)
        all_rows.append([
            path,
            dept,
            item.get("name", ""),
            size_bytes,
            vn_time,
            deleted_by
        ])

        total_items += 1

    NEXT_URL = data.get("@odata.nextLink")


# =========================
# SORT (newest first)
# =========================

all_rows.sort(key=lambda x: x[4], reverse=True)

# =========================
# EXCEL
# =========================

wb = Workbook()
wb.remove(wb.active)
header_font = Font(bold=True)

# =========================
# SHEET: RECYCLE BIN
# =========================

ws = wb.create_sheet("RECYCLE_BIN")

ws.append([
    "deletedFromLocation",
    "Dept",
    "name",
    "size_bytes",
    "deletedDateTime_VN",
    "deletedBy"
])

for c in ws[1]:
    c.font = header_font

for r in all_rows:
    ws.append(r)

# auto width simple (fast)
for col in ws.columns:
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 30


# =========================
# DASHBOARD
# =========================

dashboard = wb.create_sheet("DASHBOARD")

dashboard.append(["Dept", "Files", "Total Size"])

for c in dashboard[1]:
    c.font = header_font

sorted_depts = sorted(dept_total_size.items(), key=lambda x: x[1], reverse=True)

for dept, size_bytes in sorted_depts:
    dashboard.append([
        dept,
        dept_file_count.get(dept, 0),
        format_size(size_bytes)
    ])

dashboard.append([])
dashboard.append([
    "TOTAL",
    total_items,
    format_size(sum(dept_total_size.values()))
])

# =========================
# SAVE
# =========================

output_file = "recyclebin_report.xlsx"
wb.save(output_file)

print("\n==============================")
print(f"Pages processed: {page_count}")
print(f"Total items: {total_items}")
print(f"Output file: {output_file}")
print("==============================")
